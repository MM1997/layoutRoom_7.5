# coding=utf-8

import openpyxl
from openpyxl.styles import Font, colors, Alignment
import csv


class ReadExcle(object):
    def __init__(self,path):
        self.path = path
        self.wb = openpyxl.load_workbook(path)

    def get_column_data(self,column):
        sheet_names = self.wb.sheetnames
        sheet = self.wb[sheet_names[0]]
        max_row = sheet.max_row
        print(max_row)
        data =[]
        name = {}
        for i in range(2, max_row):
            if isinstance(sheet[column + str(i)].value,int):
                #取出列的值
                columnValue = sheet[column + str(i)].value
                #转为字符串
                columnValue = str(columnValue)
                data.append(columnValue)
        # name[sheet[column + str(1)].value] = list(set(data))
        return list(set(data))

    def get_solutionId(self):
        id= {}
        #a列 solutionId
        solutionId = self.get_column_data( "a")
        # b列 dnaSolutionId
        dnaSolutionId = self.get_column_data("b")
        id["solutionId"] = solutionId
        id["dnaSolutionId"] = dnaSolutionId
        return id



# path = r'..\test2.xlsx'
# wb = ReadExcle(path)
# id = wb.get_solutionId()
# print(id)

