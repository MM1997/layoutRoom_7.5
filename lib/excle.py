# coding=utf-8

import openpyxl
from openpyxl.styles import Font, colors, Alignment
import csv
from aijia import jsonData
from aijia import appaj

def excle1():
    try:
        #打开excle
        wb = openpyxl.load_workbook(r'..\test2.xlsx')
        #获取sheet name
        print(wb.sheetnames)

        #指定某个sheet
        sheet = wb['Sheet1']
        print(sheet)

        #获取sheet民称
        print(sheet.title)

        #活动表
        print('活动表')
        print(wb.active)

        #获取数据
        c = sheet['B1']
        print(c.value)
        print(c.row)
        print(c.column)

        #cell
        print(sheet.cell(row=1,column=2).value)
        print(sheet.cell(row=1,column=2).row)
        print(sheet.cell(row=1,column=2).column)



        #列字母和数字间的转换
        print(22222222222)
        print(sheet.max_column) #获取最大列数
        print(sheet.max_row) #获取最大行数
        print(sheet.min_row)#获取最小行数
        print(sheet.min_column) #获取最小列数
        # print sheet.rows.value
        print(sheet.dimensions) #表格的大小，这里的大小是指含有数据的表格的大小，即：左上角的坐标:右下角的坐标

        for row in sheet.iter_rows(min_row=2,max_row=4,min_col=2,max_col=4):
            print(row)

        # print tuple(sheet['A1':'C3'])

        for i in sheet['A1':'C3']:
            for j in i:
                j.value=2

        sheet['A1']='见'
        # print wb.create_sheet()
        # print wb.sheetnames

        # wb.remove_sheet(wb['Sheet'])


        #设置字体
        #设置字体样式,12号字体，加粗
        font1 = Font(name=u'宋体', size=12, italic=False,bold=True)
        sheet['A1'].font = font1

        #居中
        sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')

        # 第1行行高
        sheet.row_dimensions[1].height = 20
        # A列列宽
        sheet.column_dimensions['A'].width = 20
        wb.save(r'..\test2.xlsx')

    finally:
        wb.close()



def write_to_excle(path,hash):
    # 打开excle
    wb = openpyxl.load_workbook(path)
    # 获取sheet name
    print(wb.sheetnames)

    # 指定某个sheet
    sheet = wb['Sheet1']

    for area in hash["area"].keys():
        #写入当前时间
        print(sheet.max_column)  # 获取最大列数
        print(sheet.max_row)  # 获取最大行数
        # print(sheet.min_row)  # 获取最小行数
        # print(sheet.min_column)  # 获取最小列数
        row = sheet.max_row + 1

        #存放时间
        sheet['A'+ str(row)] = hash["time"]

        #存放产品方案id
        sheet['B' + str(row)] = int(hash["solutionId"])
        # 存放DNA方案id
        sheet['C' + str(row)] = int(hash["dnaSolutionId"])

        # 存放房间区域
        sheet['D' + str(row)] = hash["area"][area]["description"]
        # 存放房间区域下的物品数量
        sheet['E' + str(row)] = int(hash["area"][area]["num"])
        # 存放智能套用前物体个数
        sheet['F' + str(row)] = int(hash["area"][area]["data"]["before"]["objectNum"])
        # 存放智能套用后物体个数
        sheet['G' + str(row)] = int(hash["area"][area]["data"]["after"]["objectNum"])
        wb.save(r'..\test2.xlsx')
        wb.close()


path = r'..\test2.xlsx'
# hash = jsonData.saveTime("dsdd")


dic = {
    "time": "",
    "solutionId": "10011",
    "dnaSolutionId": "6008",
    "area":{
            "livingRoom": {
            "description": "客厅",
            "num":"1",
            "data":{
                "before": {"objectNum": "10", "selectNum": "10"},
                "after": {"objectNum": "11", "selectNum": "11"}
            }
        },
        "restaurant": {
            "description": "餐厅",
            "num":"2",
            "data":{
                "before": {"objectNum": "20", "selectNum": "20"},
                "after": {"objectNum": "21", "selectNum": "21"}
            }
        },
        "masterBedroom": {
            "description": "主卧",
            "num":"3",
            "data":{
                "before": {"objectNum": "30", "selectNum": "30"},
                "after": {"objectNum": "31", "selectNum": "31"}
            }
        },
        "secondBedroom": {
            "description": "次卧",
            "num":"4",
            "data": {
                "before": {"objectNum": "40", "selectNum": "40"},
                "after": {"objectNum": "41", "selectNum": "41"}
            }
        },
        "kitchen": {
            "description": "厨房",
            "num":"5",
            "data": {
                "before": {"objectNum": "50", "selectNum": "50"},
                "after": {"objectNum": "51", "selectNum": "51"}
            }
        },
        "bathroom": {
            "description": "主卫",
            "num":"6",
            "data": {
                "before": {"objectNum": "60", "selectNum": "60"},
                "after": {"objectNum": "61", "selectNum": "61"}
            }
        }
    }
}

dic["time"] = appaj.getCurrentTime()

write_to_excle(path,dic)




