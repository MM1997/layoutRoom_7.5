# coding=utf-8
import openpyxl
from openpyxl.styles import Font, colors, Alignment,Border,Side
from openpyxl.utils import get_column_letter,column_index_from_string
from statistical.readMysql import Mysql,generate_new_dict
from openpyxl import Workbook
import pandas as pd
from pandas import DataFrame
import time
import os

class Excle(object):
    def __init__(self,path):
        self.path = path
        if os.path.exists(self.path): #判断是否存在excle表，不存在就自动创建一个
            self.wb = openpyxl.load_workbook(path)
            sheet_names = self.wb.sheetnames
            self.sheet = self.wb[sheet_names[0]]

            # 创建一个新的sheet
            if not "Sheet2" in self.wb.sheetnames:
                self.wb.create_sheet("Sheet2")
            self.sheet2 = self.wb["Sheet2"]
        else:
            self.wb = Workbook()
            sheet_names = self.wb.sheetnames
            self.sheet = self.wb[sheet_names[0]]

            # 创建一个新的sheet
            if not "Sheet2" in self.wb.sheetnames:
                self.wb.create_sheet("Sheet2")
            self.sheet2 = self.wb["Sheet2"]

        #清空excle数据
        for i in  range(10):
            for col in range(1,self.sheet.max_column+1):
                self.sheet.delete_cols(col, 1)
        for i in  range(10):
            for col in range(1,self.sheet2.max_column+1):
                self.sheet2.delete_cols(col, 1)


    def initialization(self):
        """
        初始化表格,sheet1、sheet2
        :return:
        """
        #设置列宽
        sheet1_max_column = self.sheet.max_column
        sheet2_max_column = self.sheet2.max_column
        sheet1_max_row = self.sheet.max_row
        sheet2_max_row = self.sheet2.max_row
        for column in range(1,sheet1_max_column+1):
            cell = get_column_letter(column)
            self.sheet.column_dimensions[cell].width = 15
        for column in range(1,sheet2_max_column+1):
            cell = get_column_letter(column)
            self.sheet2.column_dimensions[cell].width = 15

        #居中
        for column in range(1, sheet1_max_column + 1):
            cell = get_column_letter(column)
            for row in range(1,500):
                self.sheet[cell + str(row)].alignment = Alignment(horizontal='center', vertical='center')
                # 对第一行字体加粗
                self.sheet[cell + str(1)].font = Font(name=u'宋体', size=12, italic=False, bold=True)

        for column in range(1, sheet2_max_column + 1):
            cell = get_column_letter(column)
            for row in range(1,500):
                self.sheet2[cell + str(row)].alignment = Alignment(horizontal='center', vertical='center')
                self.sheet2[cell + str(1)].font = Font(name=u'宋体', size=12, italic=False, bold=True)
                #设置sheet2第一行的背景色
                self.sheet2[cell + str(1)].font = Font('宋体', size=12, color='000000FF', bold=True)
                # 设置sheet2第一行的字体颜色
                self.sheet2[cell + str(1)].fill = openpyxl.styles.fills.GradientFill(stop=['00CCFF'])


        # 设置sheet2的边框
        border = Border(left=Side(style='medium', color='FF000000'),
                        right=Side(style='medium', color='FF000000'),
                        top=Side(style='medium', color='FF000000'),
                        bottom=Side(style='medium', color='FF000000'),
                        diagonal=Side(style='medium', color='FF000000'), diagonal_direction=0,
                        outline=Side(style='medium', color='FF000000'),
                        vertical=Side(style='medium', color='FF000000'),
                        horizontal=Side(style='medium', color='FF000000'))
        for culoum in range(1,sheet2_max_column+1):
            for row in range(1, sheet2_max_row+1):
                self.sheet2.cell(row=row, column=culoum).border = border



    def write_to_excel(self,info,condition):
        """
        :param info: 字典格式，sql查询结果
        :param condition: sql查询的条件，如：根据sr_key_x条件查询，condition可写成condition=[0.85,0.9,0.95]
        :return:
        """
        max_row = self.sheet.max_row #获取最大行数

        start_column = 2 #定义起始的列数
        dict = {} #定义一个字典，用来索引每个字段的列数
        for key in info[0].keys():
            dict[key] = start_column
            start_column += 1
        dict["condition"] = 1

        #写入第一列
        for key in dict:
            self.sheet.cell(row=1, column=dict[key]).value = key

        #将数据库查询的结果写入到excle
        for info1 in info:
            max_row = self.sheet.max_row  # 获取最大行数
            for key in info1.keys():
                self.sheet.cell(row=max_row + 1, column=1).value = condition
                self.sheet.cell(row=max_row+1, column=dict[key]).value = info1[key]


        self.wb.save(self.path)
        self.wb.close()

    def columns_num(self,conditions):
        """
         定义一个字典用来存储统计报表的各个字段的列数，如：
         {'85%占比': 3, '90%占比': 4, '评分统计': 1, '相似度统计': 2, '总数': 5}
         '85%占比'是在第3列
        :param conditions: 查询的条件，如：根据sr_key_x条件查询，condition可写成condition=[0.85,0.9,0.95]
        :return:
        """
        columns = {}
        # 第一列
        self.sheet2.cell(row=1, column=1).value = "评分统计"
        self.sheet2.cell(row=1, column=2).value = "相似度统计"
        for condition in conditions[:-1]:
            max_column = self.sheet2.max_column
            self.sheet2.cell(row=1, column=max_column + 1).value = str(int(condition * 100)) + "%占比"
            columns[str(int(condition * 100)) + "%占比"] = max_column + 1
        self.sheet2.cell(row=1, column=self.sheet2.max_column + 1).value = "总数"
        columns["评分统计"] = 1
        columns["相似度统计"] = 2
        columns["总数"] = self.sheet2.max_column
        # print(columns)
        return columns

    # excle.statistical(datas, keys=["room_name"], values=["count(1)"], conditions=sr_key_xs)

    def statistical(self,info,keys=["room_name","solutionId"],values=["count"],conditions=[0.85,0.9,1]):
        """

        :param info1: list格式，是根据条件查询出的数据集,最后一个值作为参考值，info[i]/info[-1]
        # :param info2: 参考数据，也是根据条件查询出的数据，info1/info2
        :param keys: 需要生成的新的key值
        :param values: 以values为key生成的value值，和keys组成生成新的dict
        :param condition:sql查询的条件，如：根据sr_key_x条件查询，condition可写成condition=[0.85,0.9,0.95]
        :return:
        room_name          solutionId          count
        1                       2               3
        2                       3               4

        room_name          solutionId          count
        1                       2               2
        2                       3               1
        3                       1               1
        """



        if len(info) != len(conditions):
            raise "info1和condition长度必须相同，请检查！"

        #首先将sql语句查询到的数据写到sheet1里
        for i in range(len(info)):
            self.write_to_excel(info[i], conditions[i])

        #定义一个字典用来存储统计报表的各个字段的列数,
        # {'85%占比': 3, '90%占比': 4, '评分统计': 1, '相似度统计': 2, '总数': 5}
        #'85%占比'是在第3列
        columns = self.columns_num(conditions)


        print(keys)
        print(values[0])
        #info最后一个keys值,作为参考
        last_info_keys = []
        last_info = generate_new_dict(info[-1], keys=keys, values=values[0])
        for k in last_info:
            last_info_keys.append(k)
        print("last_info_keys:",last_info_keys)


        #[{'相似度统计': '主卧', '85%占比': '87.20'}, {'相似度统计': '书房', '85%占比': '1.59'}, {'相似度统计': '主卧', '90%占比': '35.05'}]
        num = 0
        datas = []
        for info1 in info[:-1]:
            new_info = generate_new_dict(info1, keys=keys, values=values[0])
            for key in last_info_keys:
                data = {}
                data["相似度统计"] = key
                data[str(int(conditions[num]*100))+"%占比"] = "{:.2f}".format(100*(1-new_info.get(key,0)/last_info.get(key)))
                datas.append(data)
            num += 1
        # print("datas:",datas)

        #将datas里的"相似度统计"对应的value值相同的dict合并
        #[{'相似度统计': '主卧', '85%占比': '87.20', '90%占比': '74.51'}, {'相似度统计': '书房', '85%占比': '1.59', '90%占比': '1.59'}]
        dict_datas = []
        for key in last_info_keys:
            dict = {}
            for info in datas:
                if info["相似度统计"] == key:
                    dict.update(info)
            dict_datas.append(dict)
        print("dict_datas:",dict_datas)

        #写入sheet2里，进行报表统计
        row = 2
        max_column = self.sheet2.max_column
        for info in dict_datas:
            for key in info.keys():
                self.sheet2.cell(row=row, column=columns[key]).value = info[key]
                self.sheet2.cell(row=row, column=max_column).value = last_info[info["相似度统计"]]
            row += 1

        #excle的单元格处理
        self.initialization()

        self.wb.save(self.path)
        self.wb.close()

path = r'..\statistical.xlsx'
# Excle(path)