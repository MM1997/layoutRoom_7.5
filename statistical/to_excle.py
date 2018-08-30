import openpyxl
from openpyxl.styles import Font, colors, Alignment,Border,Side
from openpyxl.utils import get_column_letter,column_index_from_string
from statistical.readMysql import Mysql
import time

class Excle(object):
    def __init__(self,path,sr_key_x):
        self.path = path
        self.sr_key_x = str(sr_key_x)
        self.wb = openpyxl.load_workbook(path)
        sheet_names = self.wb.sheetnames
        self.sheet = self.wb[sheet_names[0]]
        # print(self.wb.active)

    def initialization(self):
        """
        初始化表格
        :return:
        """
        #设置列宽
        self.sheet.column_dimensions['A'].width = 15
        #将首行进行合并单元格
        self.sheet.merge_cells('A1:D1')
        self.sheet.merge_cells('F1:I1')
        self.sheet.merge_cells('K1:N1')
        self.sheet.merge_cells('P1:S1')
        #将首行赋值
        self.sheet["A1"] = 0.85
        self.sheet["F1"] = 0.9
        self.sheet["K1"] = 0.95
        self.sheet["P1"] = 1
        #居中
        column = column_index_from_string("S")
        for num in range(1,column+1): #["A","B","C","D","E","F","G","H","I","J","K","l","M","N","O","P","Q","R","S"]:
            #将数字转化字母
            cell = get_column_letter(num)
            for i in range(1,30):
                self.sheet[cell+str(i)].alignment = Alignment(horizontal='center', vertical='center')
                #对第一、二行字体加粗
                if i < 3:
                    # font = Font('黑体', bold=True)
                    font = Font(name=u'宋体', size=12, italic=False, bold=True)
                    self.sheet[cell+str(i)].font = font

        self.wb.save(self.path)

    def statistical_data(self):
        """
        评分统计，从20行开始写入数据并统计
        :return:
        """
        #20行写入数据
        datas = ["评分统计","相似度统计","85%占比","90%占比","95%占比","总数"]
        culoums = ["A", "B", "C", "D", "E","F"]
        #设置边框
        border = Border(left=Side(style='medium', color='FF000000'), right=Side(style='medium', color='FF000000'),
                        top=Side(style='medium', color='FF000000'), bottom=Side(style='medium', color='FF000000'),
                        diagonal=Side(style='medium', color='FF000000'), diagonal_direction=0,
                        outline=Side(style='medium', color='FF000000'), vertical=Side(style='medium', color='FF000000'),
                        horizontal=Side(style='medium', color='FF000000'))
        #设置第20行标题
        for i in range(len(culoums)):
            self.sheet[culoums[i]+"20"] = datas[i]
            #设置字体为蓝色
            self.sheet[culoums[i]+"20"].font = Font('宋体', size=12,color='000000FF',bold=True)
            #设置20行背景色
            self.sheet[culoums[i] + "20"].fill = openpyxl.styles.fills.GradientFill(stop=['00CCFF'])

        #设置统计评分下面的边框
        for culoum in culoums:
            for row in range(20,29):
                self.sheet[culoum + str(row)].border = border

        #相似度统计列
        areas = ["主卧","书房","儿童房","客厅","榻榻米房","次卧","老人房","餐厅"]
        num = 0
        for row in range(21,29):
            self.sheet["B" + str(row)] = areas[num]
            num += 1

        #85%占比列
        self.sheet["C21"] = "=100-100*D3/S3"
        self.sheet["C22"] = "=100-100*D4/S4"
        for i in range(1,9):
            j = i + 2
            self.sheet["C2" + str(i)] = "=100-100*D{}/S{}".format(j,j)
            # 设置小数点位数为2位
            self.sheet["C2" + str(i)].number_format = '0.00'
        # 90%占比列
        for i in range(1, 9):
            j = i + 2
            self.sheet["D2" + str(i)] = "=100-100*I{}/S{}".format(j, j)
            #设置小数点位数为2位
            self.sheet["D2" + str(i)].number_format = '0.00'

        # 95%占比列
        for i in range(1, 9):
            j = i + 2
            self.sheet["E2" + str(i)] = "=100-100*N{}/S{}".format(j, j)
            # 设置小数点位数为2位
            self.sheet["E2" + str(i)].number_format = '0.00'
        self.wb.save(self.path)
        # 总数比列
        for i in range(1, 9):
            j = i + 2
            self.sheet["F2" + str(i)] = "=S{}".format(j)
            # 设置小数点位数为2位
            self.sheet["F2" + str(i)].number_format = '0.00'
        self.wb.save(self.path)

    def row_column(self,data,row,column):
        """

        :param data: 从数据库获取的数据 【字典】
        :param row: 行
        :param column: 列
        :return:
        """
        self.sheet.cell(row=2, column=column).value = "room_name"
        self.sheet.cell(row=2, column=column + 1).value = "solutionId"
        self.sheet.cell(row=2, column=column + 2).value = "dnaSolutionId"
        self.sheet.cell(row=2, column=column + 3).value = "count(1)"

        self.sheet.cell(row=row, column=column).value = data.setdefault("room_name", "")
        self.sheet.cell(row=row, column=column + 1).value = data.setdefault("solutionId","")
        self.sheet.cell(row=row, column=column + 2).value = data.setdefault("dnaSolutionId","")
        self.sheet.cell(row=row, column=column + 3).value = data.setdefault("count(1)","")


    def to_excle(self,datas):
        """
        将从数据库里获取的数据写到excle里
        :param datas: 从数据库获取的数据 【字典】
        :return:
        """
        #初始化表格
        self.initialization()

        if self.sr_key_x == "0.85":
            columnNmae = "a"
        elif self.sr_key_x == "0.9":
            columnNmae = "f"
        elif self.sr_key_x == "0.95":
            columnNmae = "k"
        elif self.sr_key_x == "1":
            columnNmae = "p"

        for data in datas:
            column = column_index_from_string(columnNmae)
            # print(column)
            if data["room_name"]  == "主卧":
                self.row_column(data,3,column)
            elif data["room_name"]  == "书房":
                self.row_column(data, 4, column)
            elif data["room_name"]  == "儿童房":
                self.row_column(data, 5, column)
            elif data["room_name"]  == "客厅":
                self.row_column(data, 6, column)
            elif data["room_name"]  == "榻榻米房":
                self.row_column(data, 7, column)
            elif data["room_name"]  == "次卧":
                self.row_column(data, 8, column)
            elif data["room_name"]  == "老人房":
                self.row_column(data, 9, column)
            elif data["room_name"]  == "餐厅":
                self.row_column(data, 10, column)
        self.wb.save(self.path)
        self.wb.close()


def CalcTwoResult(info1,info2,keys=["aa","bb"],values=["cc"],fun="={}/{}"):
    '''

    aa          bb          cc
    1           2           3
    2           3           4

    aa          bb          cc
    1           2           2
    2           3           1
    3           1           1
    '''
    pass


if __name__ == "__main__":
    t =  time.time()
    path = r'statistical.xlsx'
    mysql = Mysql()
    sr_key_xs = [0.85,0.9,0.95,1]
    for sr_key_x in sr_key_xs:
        sql = 'select room_name,solutionId,dnaSolutionId ,count(1) from data.layout_best_infos where 1=1 ' \
              'and sr_key_x>0 ' \
              'and sr_key_x<{} ' \
              'and room_name in ("客厅","餐厅","主卧","次卧","儿童房","老人房","书房","榻榻米房") ' \
              'group by room_name ' \
              'order by room_name limit 100'.format(sr_key_x)

        data = mysql.get_data(sql)
        #将数据库里取到的值存到excle里
        excle = Excle(path,sr_key_x)
        excle.to_excle(data)
        #评分统计
        excle.statistical_data()

    print(time.time()-t)